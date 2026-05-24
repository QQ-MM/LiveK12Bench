"""Average metrics across all paper-sheets in a metrics workbook.

Reads every per-paper sheet in the input Excel file (skipping the summary
sheet itself) and writes a 'summary' sheet containing per-model averages
of every numeric column. Fractional `a/b` cells are averaged separately as
sums of numerators and denominators, then re-rendered as `mean_num/mean_den`.

Usage:
    python evaluate/util/average_metrics.py --xlsx path/to/metrics.xlsx \\
        [--summary-sheet 汇总]
"""

import argparse
import os

import numpy as np
import pandas as pd
from openpyxl import load_workbook


def split_frac(x):
    if isinstance(x, str) and "/" in x:
        a, b = x.split("/")
        return float(a), float(b)
    return (np.nan, np.nan)


def average(xlsx_path, summary_sheet="汇总"):
    xl = pd.ExcelFile(xlsx_path)
    sheet_names = [s for s in xl.sheet_names if s != summary_sheet]
    if not sheet_names:
        raise ValueError(f"No paper sheets found in {xlsx_path} (only {xl.sheet_names}).")
    dfs = {s: xl.parse(s, header=[0, 1], index_col=0) for s in sheet_names}

    sample_df = next(iter(dfs.values()))
    fract_cols = sample_df.columns[-2:]

    numeric_acc = None
    fraction_num = None
    fraction_den = None
    cnt = len(dfs)

    for df in dfs.values():
        num_df = pd.DataFrame(
            {col: df[col].map(lambda x: split_frac(x)[0]) for col in fract_cols},
            index=df.index,
        ).astype(float)
        den_df = pd.DataFrame(
            {col: df[col].map(lambda x: split_frac(x)[1]) for col in fract_cols},
            index=df.index,
        ).astype(float)
        df_numeric = df.drop(columns=fract_cols).apply(pd.to_numeric, errors="coerce")

        numeric_acc = df_numeric if numeric_acc is None else numeric_acc.add(df_numeric, fill_value=0)
        fraction_num = num_df if fraction_num is None else fraction_num.add(num_df, fill_value=0)
        fraction_den = den_df if fraction_den is None else fraction_den.add(den_df, fill_value=0)

    numeric_avg = numeric_acc / cnt
    numerator_avg = fraction_num / cnt
    denominator_avg = fraction_den / cnt

    frac_avg = numerator_avg.round(2).astype(str) + "/" + denominator_avg.round(2).astype(str)
    frac_avg.columns = fract_cols

    summary_df = pd.concat([numeric_avg, frac_avg], axis=1)

    # Find a "总分" column (final-score) and sort rows by it descending.
    possible = [
        col for col in summary_df.columns
        if (isinstance(col, tuple) and col[-1] == "总分") or (not isinstance(col, tuple) and col == "总分")
    ]
    if not possible:
        raise ValueError("Could not locate the '总分' column in any sheet.")
    score_col = possible[0]
    summary_df = summary_df.sort_values(score_col, ascending=False)

    # Replace any existing summary sheet, then move it to the front.
    wb = load_workbook(xlsx_path)
    if summary_sheet in wb.sheetnames:
        del wb[summary_sheet]
    wb.save(xlsx_path)

    with pd.ExcelWriter(
        xlsx_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
    ) as writer:
        summary_df.to_excel(writer, sheet_name=summary_sheet)

    wb = load_workbook(xlsx_path)
    ws = wb[summary_sheet]
    wb._sheets.remove(ws)
    wb._sheets.insert(0, ws)
    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", required=True, help="Path to the metrics workbook.")
    parser.add_argument("--summary-sheet", default="汇总",
                        help="Name of the summary sheet (will be replaced).")
    args = parser.parse_args()

    if not os.path.exists(args.xlsx):
        parser.error(f"File not found: {args.xlsx}")

    average(args.xlsx, summary_sheet=args.summary_sheet)
    print("Done!")


if __name__ == "__main__":
    main()
